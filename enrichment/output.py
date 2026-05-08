import io
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASS, EMAIL_FROM

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COLUMNS = [
    ("Firma",                20),
    ("Sitz",                 15),
    ("Branche",              20),
    ("Mitarbeiter",          12),
    ("Umsatz",               15),
    ("Verband",              20),
    ("Summary",              45),
    ("Pain Point 1",         40),
    ("Pain Point 2",         40),
    ("Pain Point 3",         40),
    ("Investitionssignale",  40),
    ("MH: Economic Buyer",   30),
    ("MH: Gesprächseinstieg",45),
    ("MH: Talking Points",   50),
    ("MH: Red Flags",        35),
    ("ZEFIX UID",            18),
    ("ZEFIX URL",            35),
    ("Domain",               25),
    ("Stand",                18),
]


def results_to_excel(results: list[dict], existing_wb_bytes: bytes | None = None) -> bytes:
    """
    Create (or update) an Excel workbook with enrichment results.
    Returns the workbook as bytes for download.
    """
    if existing_wb_bytes:
        wb = openpyxl.load_workbook(io.BytesIO(existing_wb_bytes))
        ws = wb.active
        _ensure_headers(ws)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Firmen Enrichment"
        _write_headers(ws)

    for i, r in enumerate(results, start=2):
        _write_row(ws, i, r)
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_headers(ws):
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22


def _ensure_headers(ws):
    if ws.cell(row=1, column=1).value != "Firma":
        _write_headers(ws)


def _write_row(ws, row: int, r: dict):
    meta = r.get("_meta", {})
    signals = "; ".join(r.get("investitionssignale", []))
    mh = r.get("miller_heiman", {})
    bi = mh.get("buying_influences", {}) if mh else {}
    tps = "; ".join(mh.get("talking_points", [])) if mh else ""
    flags = "; ".join(mh.get("red_flags", [])) if mh else ""

    values = [
        r.get("company", meta.get("companyName", "")),
        r.get("sitz", ""),
        r.get("branche", ""),
        r.get("mitarbeiter", ""),
        r.get("umsatz", ""),
        r.get("verband", ""),
        r.get("summary", ""),
        r.get("pain_point_1", ""),
        r.get("pain_point_2", ""),
        r.get("pain_point_3", ""),
        signals,
        bi.get("economic_buyer", ""),
        mh.get("recommended_entry", "") if mh else "",
        tps,
        flags,
        r.get("uid", ""),
        r.get("zefix_url", "") or (meta.get("zefix_url", "")),
        meta.get("domain", ""),
        meta.get("timestamp", "")[:10],
    ]
    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=row, column=col_idx, value=val or "")


def send_email(
    to_addresses: list[str],
    results: list[dict],
    excel_bytes: bytes,
    subject: str | None = None,
) -> None:
    """Send enrichment results by email with Excel attachment."""
    subject = subject or f"Firmen-Enrichment – {datetime.now().strftime('%d.%m.%Y')}"

    body_lines = [
        f"<h2>Firmen-Enrichment Report</h2>",
        f"<p>Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>",
        "<table border='1' cellpadding='6' cellspacing='0' style='border-collapse:collapse;font-family:Calibri;font-size:12px;'>",
        "<tr style='background:#1F3864;color:white;'>"
        "<th>Firma</th><th>Pain Point 1</th><th>Pain Point 2</th><th>Pain Point 3</th></tr>",
    ]
    for i, r in enumerate(results):
        bg = "#EBF3FB" if i % 2 == 0 else "#FFFFFF"
        body_lines.append(
            f"<tr style='background:{bg};'>"
            f"<td><b>{r.get('company','')}</b><br><small>{r.get('summary','')}</small></td>"
            f"<td>{r.get('pain_point_1','')}</td>"
            f"<td>{r.get('pain_point_2','')}</td>"
            f"<td>{r.get('pain_point_3','')}</td></tr>"
        )
    body_lines.append("</table><p>Details im angehängten Excel.</p>")

    msg = MIMEMultipart()
    msg["From"] = EMAIL_FROM
    msg["To"] = ", ".join(to_addresses)
    msg["Subject"] = subject
    msg.attach(MIMEText("\n".join(body_lines), "html", "utf-8"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    filename = f"Enrichment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(EMAIL_FROM, to_addresses, msg.as_string())
